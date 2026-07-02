import sys
sys.path.append('./')
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from biomarker.analysis.figure_4c import generate_osa_figure
from biomarker.analysis.figure_4a import generate_other_medications_figure
from biomarker.analysis.figure_4e import generate_mdd_confound_figure
from biomarker.analysis.figure_4d import generate_fairness_analysis_figure, load_data as load_bmi_data
from biomarker.analysis.figure_4b import generate_cotherapy_analysis_figure

font_size = 16
def reset_font_size():
    plt.rcParams.update({
        "font.size": font_size,
        "axes.titlesize": font_size,
        "axes.labelsize": font_size,
        "xtick.labelsize": font_size,
        "ytick.labelsize": font_size,
        "legend.fontsize": font_size,
        "figure.titlesize": font_size,
    })

## create a gridspec with 3 elements in top row and 2 elements in bottom row
gs = gridspec.GridSpec(3, 4, figure=plt.figure(figsize=(13, 13)))

## create a figure with the gridspec
fig = plt.figure(figsize=(13, 13))

## create 3 subplots in the top row (each spanning 2 columns)
ax00 = fig.add_subplot(gs[0, 0:2])  # top left
ax01 = fig.add_subplot(gs[0, 2:4])  # top middle
ax02 = fig.add_subplot(gs[1, 0:2])  # top right

## create 2 subplots in the bottom row (each spanning 3 columns)
ax10 = fig.add_subplot(gs[1, 2:4])  # bottom left
ax10.set_xlim(ax10.get_xlim()[0] - 1, ax10.get_xlim()[1])
ax11 = fig.add_subplot(gs[2, :])  # bottom right

reset_font_size()

print('Generating mdd confound figure...')
ax11, data_4e = generate_mdd_confound_figure(ax=ax11, save=False)
ax11.set_title('Robustness across Depression Severity Levels', fontsize=font_size, pad=22)

reset_font_size()
print('Generating fairness analysis figure...')
generate_fairness_analysis_figure(ax=ax10, save=False)
ax10.set_title('Robustness across BMI Levels', fontsize=font_size, pad=22)

reset_font_size()
print('Generating cotherapy analysis figure...')
ax01, data_4b = generate_cotherapy_analysis_figure(ax=ax01, save=False)
ax01.set_title('Robustness to Drug Co-Therapy', fontsize=font_size, pad=12)

reset_font_size()
print('Generating other medications figure...')
ax00, data_4a = generate_other_medications_figure(ax=ax00, save=False)
ax00.set_title('Robustness to Psychotropic & Anticholinergic Medications', fontsize=font_size, pad=12)

reset_font_size()
print('Generating osa figure...')
ax02, data_4c = generate_osa_figure(ax=ax02, save=False)
ax02.set_title('Robustness to Sleep Apnea', fontsize=font_size, pad=22)

plt.tight_layout()
plt.subplots_adjust(hspace=0.58, wspace=0.35)
plt.savefig('biomarker/analysis/figure_4.pdf', bbox_inches='tight')
print('Saved figure_4.pdf')

# --- collect BMI data for xlsx (figure_4d) ---
_df_bmi = load_bmi_data()
_df_bmi = _df_bmi[_df_bmi['mit_bmi'].notna()].copy()
_bmi_labels = ['<18.5', '18.5-25', '25-30', '30-35', '35-40', '>40']
_df_bmi['bmi_bin'] = pd.cut(_df_bmi['mit_bmi'], bins=[0, 18.5, 25, 30, 35, 40, 100],
                             labels=_bmi_labels, include_lowest=True)
_df_bmi['Group'] = _df_bmi['label'].apply(lambda x: 'Control' if x == 0 else 'Antidepressant')
data_4d = _df_bmi[['bmi_bin', 'Group', 'pred']].copy()


def write_source_data_xlsx(data_4a, data_4b, data_4c, data_4d, data_4e):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUB_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUB_FONT    = Font(bold=True)

    wb = Workbook()

    def write_single_groups(ws, df, group_col, value_col, group_order):
        col = 1
        for group in group_order:
            vals = df[df[group_col] == group][value_col].round(3).values
            if len(vals) == 0:
                continue
            cell = ws.cell(row=1, column=col, value=str(group).replace('\n', ' '))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            for ri, v in enumerate(vals):
                ws.cell(row=2 + ri, column=col, value=float(v))
            ws.column_dimensions[get_column_letter(col)].width = 20
            col += 2  # blank separator

    def write_paired_groups(ws, df, group_col, hue_col, value_col, group_order, hue_order):
        col = 1
        for group in group_order:
            cell = ws.cell(row=1, column=col, value=str(group))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + len(hue_order) - 1)
            for j, hue in enumerate(hue_order):
                mask = (df[group_col].astype(str) == str(group)) & (df[hue_col] == hue)
                vals = df[mask][value_col].round(3).values
                c = ws.cell(row=2, column=col + j, value=str(hue))
                c.font = SUB_FONT
                c.fill = SUB_FILL
                c.alignment = Alignment(horizontal='center')
                for ri, v in enumerate(vals):
                    ws.cell(row=3 + ri, column=col + j, value=float(v))
                ws.column_dimensions[get_column_letter(col + j)].width = 18
            col += len(hue_order) + 1  # blank separator

    # Sheet 1: Other Medications (4a)
    ws1 = wb.active
    ws1.title = "4a Other Medications"
    order_4a = ['No Psycho\ntropic', 'Anti-\ncholinergics', 'Hypnotics', 'Anti-\nconvulsants',
                'Benzo-\ndiazepines', 'Anti-\npsychotics', 'Anti-\ndepressants']
    write_single_groups(ws1, data_4a, 'medication', 'pred', order_4a)

    # Sheet 2: Co-therapy (4b)
    ws2 = wb.create_sheet("4b Co-therapy")
    order_4b = ['Controls', 'Single\nAntidep', 'Antidep+\nAnti-\nconvulsant',
                'Antidep+\nHypnotic', 'Antidep+\nBenzo-\ndiazepine',
                'Antidep+\nAnti-\npsychotic', 'Multi-\nAntidep']
    write_single_groups(ws2, data_4b, 'cohort', 'pred', order_4b)

    # Sheet 3: OSA (4c)
    ws3 = wb.create_sheet("4c OSA")
    order_4c = ['Normal', 'Mild OSA', 'Moderate OSA', 'Severe OSA']
    write_paired_groups(ws3, data_4c, 'osa_group', 'Group', 'pred',
                        order_4c, ['Control', 'Antidepressant'])

    # Sheet 4: BMI (4d)
    ws4 = wb.create_sheet("4d BMI")
    order_4d = ['<18.5', '18.5-25', '25-30', '30-35', '35-40', '>40']
    write_paired_groups(ws4, data_4d, 'bmi_bin', 'Group', 'pred',
                        order_4d, ['Control', 'Antidepressant'])

    # Sheet 5: MDD Severity (4e)
    ws5 = wb.create_sheet("4e MDD Severity")
    order_4e = ['<30', '30-40', '40-50', '50-60', '60+']
    write_paired_groups(ws5, data_4e, 'zung_index_bin', 'group', 'pred',
                        order_4e, ['Control', 'Antidepressant'])

    out_path = 'biomarker/analysis/figure_4_source_data.xlsx'
    wb.save(out_path)
    print(f'Saved {out_path}')


write_source_data_xlsx(data_4a, data_4b, data_4c, data_4d, data_4e)
print('Done')
