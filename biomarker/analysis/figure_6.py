import os
import random
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import scipy.stats
from scipy.ndimage import gaussian_filter
from scipy.spatial.distance import pdist, squareform
from scipy.stats import pearsonr as _pearsonr
from sklearn.manifold import TSNE
from tqdm import tqdm

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../data')
OUT_DIR  = os.path.dirname(os.path.abspath(__file__))

FONT_SIZE = 9


# ── Shared utilities ─────────────────────────────────────────────────────────

def pearsonr_nan(x, y, return_pval=False):
    nans = np.isnan(x) | np.isnan(y)
    x, y = x[~nans], y[~nans]
    if not return_pval:
        return np.round(scipy.stats.pearsonr(x, y).correlation, 3)
    return (np.round(scipy.stats.pearsonr(x, y).correlation, 3),
            float(f"{scipy.stats.pearsonr(x, y).pvalue:.0e}"))

def smooth(arr, n):
    pad_width = n // 2
    padded = np.pad(arr, pad_width, mode='edge')
    return np.convolve(padded, np.ones(n) / n, mode='valid')


# ── Figure 6a/b helpers ───────────────────────────────────────────────────────

def tsne_progression_plot(X, y, method='points', grid_size=200, point_frac=0.2,
                           min_count=3, smooth_sigma=1.2, cmap='viridis',
                           overlay_points_frac=0.02, s=2, alpha=0.25,
                           contour=False, contour_levels=10, ax=None,
                           random_state=0, rasterized=True, ticks=None, label=None):
    X = np.asarray(X, dtype=float)
    y = np.asarray(y, dtype=float)
    y_min_p, y_max_p = np.percentile(y, 5), np.percentile(y, 95)
    y = np.clip(y, y_min_p, y_max_p)

    rng = np.random.default_rng(random_state)
    if ax is None:
        _, ax = plt.subplots(figsize=(5, 5), dpi=150)

    x_min, x_max = X[:, 0].min(), X[:, 0].max()
    y_min, y_max = X[:, 1].min(), X[:, 1].max()
    extent = (x_min, x_max, y_min, y_max)
    im = None

    if method == 'points':
        n = X.shape[0]
        k = max(1, int(point_frac * n))
        idx = rng.choice(n, size=k, replace=False)
        ax.scatter(X[idx, 0], X[idx, 1], c=y[idx], s=s,
                   alpha=max(0.05, alpha), cmap=cmap, linewidths=0, rasterized=rasterized)
        sm = plt.cm.ScalarMappable(
            norm=plt.Normalize(vmin=y.min(), vmax=y.max()), cmap=cmap)
        plt.colorbar(sm, ax=ax, fraction=0.046, pad=0.04, label=label, ticks=ticks)
    else:
        gx = np.linspace(x_min, x_max, grid_size + 1)
        gy = np.linspace(y_min, y_max, grid_size + 1)
        H,  _, _ = np.histogram2d(X[:, 0], X[:, 1], bins=[gx, gy])
        Hw, _, _ = np.histogram2d(X[:, 0], X[:, 1], bins=[gx, gy], weights=y)

        if method == 'grid':
            with np.errstate(invalid='ignore', divide='ignore'):
                mean_grid = np.where(H >= min_count, Hw / H, np.nan)
            im = ax.imshow(np.flipud(mean_grid.T), extent=extent, origin='lower',
                           aspect='equal', cmap=cmap, interpolation='nearest',
                           rasterized=rasterized)
        elif method == 'smooth':
            Hs  = gaussian_filter(H,  smooth_sigma, mode='constant')
            Hws = gaussian_filter(Hw, smooth_sigma, mode='constant')
            with np.errstate(invalid='ignore', divide='ignore'):
                sg = np.where(Hs > 1e-6, Hws / Hs, np.nan)
            im = ax.imshow(np.flipud(sg.T), extent=extent, origin='lower',
                           aspect='equal', cmap=cmap, interpolation='bilinear',
                           rasterized=rasterized)

        if overlay_points_frac and overlay_points_frac > 0:
            n = X.shape[0]
            k = max(1, int(overlay_points_frac * n))
            idx = rng.choice(n, size=k, replace=False)
            ax.scatter(X[idx, 0], X[idx, 1], s=s, c='k', alpha=0.15,
                       linewidths=0, rasterized=rasterized)

    ax.set_xticks([]); ax.set_yticks([])
    ax.set_xlim(x_min, x_max); ax.set_ylim(y_min, y_max)
    ax.set_aspect('equal', adjustable='datalim')
    if im is not None:
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label='', ticks=ticks)
    return ax, im


def plot_binned_boxplots(x, y, min_x, max_x, n_bins=8, ax=None):
    x, y = np.array(x), np.array(y)
    bins = np.linspace(min_x, max_x, n_bins)
    bins = np.insert(bins, 0, 0)
    bins = np.append(bins, 1_000_000)
    boxplots = []
    for i in range(len(bins) - 1):
        mask = (x >= bins[i]) & (x < bins[i + 1])
        boxplots.append(list(y[mask]))
    if ax is None:
        _, ax = plt.subplots(figsize=(10, 5))
    sns.boxplot(data=boxplots, ax=ax, showfliers=False, palette='Greens')
    r, pval = _pearsonr(x, y)
    pval_str = 'p<1e-10' if pval < 1e-10 else f'p = {pval:.2e}'
    ax.text(0.02, 0.8, f'Pearson r: {r:.3f}\n{pval_str}',
            ha='left', va='bottom', transform=ax.transAxes, fontsize=FONT_SIZE)
    y_pos = np.array([
        np.percentile(b, 75) if i <= 2 else np.percentile(b, 75) - 0.08
        for i, b in enumerate(boxplots)
    ])
    for i, b in enumerate(boxplots):
        ax.text(i, y_pos[i], f'N={len(b)}', ha='center', va='bottom', fontsize=FONT_SIZE - 1)
    ax.set_xticks(np.arange(len(bins) - 1))
    ax.set_xticklabels([
        f'< {bins[1]:.0f}' if i == 0
        else f'{bins[i]:.0f}+' if i == len(bins) - 2
        else f'{bins[i]:.0f}-{bins[i+1]:.0f}'
        for i in range(len(bins) - 1)
    ], fontsize=FONT_SIZE)
    ax.set_xlabel('REM Latency (min)', fontsize=FONT_SIZE)
    ax.set_ylabel('Average Model Score', fontsize=FONT_SIZE)
    ax.set_ylim(0, 1)
    return ax


def load_6ab_data():
    df = pd.read_csv(os.path.join(DATA_DIR, 'inference_v6emb_3920_all.csv'))
    df['filename'] = df['filename'].apply(lambda x: x.split('/')[-1])
    df_latency = pd.read_csv(os.path.join(DATA_DIR, 'figure_draft_v16_rem_latency.csv'))
    df_powers  = pd.read_csv(os.path.join(DATA_DIR, 'so_beta_powers_sleep_only.csv'))
    df_powers['filename'] = df_powers['concat_names'].apply(lambda x: x.split('/')[-1])
    return df, df_latency, df_powers


def get_x_y(fold, df, df_latency, df_powers, clip=True):
    df = df.copy()
    df_latency = df_latency.copy()
    df_powers  = df_powers.copy()
    df_0 = df if fold == -1 else df[df['fold'] == fold]
    df_latency = df_latency[['filename', 'rem_latency_gt']].dropna()
    df_powers  = df_powers[['filename', 'concat_beta_powers', 'concat_so_powers']].dropna()
    df_ready = df_latency.merge(df_0, on='filename', how='inner')
    df_ready = df_ready.merge(df_powers, on='filename', how='inner')
    y2 = 1 / (1 + np.exp(-df_ready['pred']))
    y3 = df_ready['label']
    if clip:
        df_ready.loc[df_ready['rem_latency_gt'] < 120, 'rem_latency_gt'] = 120
        df_ready.loc[df_ready['rem_latency_gt'] > 360, 'rem_latency_gt'] = 360
    y = df_ready['rem_latency_gt'] / 2
    x = df_ready[[c for c in df_ready.columns if c.startswith('latent_')]]
    return x, y, y2, y3, df_ready['dataset'], df_ready['concat_beta_powers'], df_ready['concat_so_powers']


PRECOMPUTED_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'figure_6_data')

def _precomputed_exists():
    return (os.path.exists(os.path.join(PRECOMPUTED_DIR, 'figure_6ab_tsne.csv')) and
            os.path.exists(os.path.join(PRECOMPUTED_DIR, 'figure_6ab_boxplot.csv')))

def compute_6ab(fold=0):
    if _precomputed_exists():
        print('[6a/b] Loading pre-computed data...')
        tsne_df = pd.read_csv(os.path.join(PRECOMPUTED_DIR, 'figure_6ab_tsne.csv'))
        box_df  = pd.read_csv(os.path.join(PRECOMPUTED_DIR, 'figure_6ab_boxplot.csv'))
        x_tsne  = tsne_df[['tsne1', 'tsne2']].values
        y_lat   = tsne_df['rem_latency_min'] / 2  # back to half-minutes to match original
        y_score = tsne_df['model_score']
        y_label = tsne_df['label']
        y_lat_all   = box_df['rem_latency_min'] / 2
        y_score_all = box_df['model_score']
        return x_tsne, y_lat, y_score, y_label, y_lat_all, y_score_all

    df, df_latency, df_powers = load_6ab_data()
    x_raw, y_all, y2_all, y3_all, _, _, _ = get_x_y(-1, df, df_latency, df_powers, clip=False)
    x, y, y2, y3, _, beta_powers, so_powers = get_x_y(fold, df, df_latency, df_powers)
    print('Computing t-SNE...')
    tsne = TSNE(n_components=2, random_state=42, perplexity=300, init='pca')
    x_tsne = tsne.fit_transform(x)
    return x_tsne, y, y2, y3, y_all, y2_all


# ── Figure 6c helpers ─────────────────────────────────────────────────────────

def process_stages(stages):
    stages = stages.copy()
    stages[stages < 0] = 0
    stages[stages > 5] = 0
    stages = stages.astype(int)
    mapping = np.array([0, 1, 2, 3, 3, 4, 0, 0, 0, 0, 0], np.int64)
    return mapping[stages]

def get_dataset(file):
    if 'cfs' in file:   return 'cfs'
    if 'mros' in file:
        if 'visit1' in file: return 'mros1'
        if 'visit2' in file: return 'mros2'
    if 'wsc' in file:   return 'wsc'
    if 'shhs1' in file: return 'shhs1'
    if 'shhs2' in file: return 'shhs2'
    return ''

ALIGNED_STATS = {
    'bwh':  {'mean': -119.816, 'std': 9.759}, 'ccshs': {'mean': -121.279, 'std': 9.403},
    'cfs':  {'mean': -121.192, 'std': 9.592}, 'chat1': {'mean': -113.508, 'std': 9.082},
    'chat2':{'mean': -114.893, 'std': 9.717}, 'chat3': {'mean': -116.877, 'std': 9.435},
    'mesa': {'mean': -122.687, 'std': 9.602}, 'mgh':   {'mean': -119.628, 'std': 9.872},
    'mgh2': {'mean': -120.553, 'std': 9.842}, 'mros1': {'mean': -120.982, 'std': 9.363},
    'mros2':{'mean': -118.224, 'std': 10.073},'p18c':  {'mean': -120.749, 'std': 10.055},
    'shhs1':{'mean': -121.102, 'std': 9.976}, 'shhs2': {'mean': -121.540, 'std': 9.873},
    'sof':  {'mean': -119.194, 'std': 10.182},'stages':{'mean': -118.392, 'std': 9.337},
    'wsc':  {'mean': -119.986, 'std': 9.550},
}

def normalize_gt(eeg, dataset):
    stats = ALIGNED_STATS[dataset]
    eeg = (eeg - stats['mean']) / stats['std'] * 7.0 + (-120.0)
    eeg = np.clip(eeg, -140, -90)
    eeg = ((eeg + 140) / 50).astype(np.float32)
    return (eeg - 0.5) / 0.5

def naive_power_post_onset(mage, stage, minutes=1_000_000, mean=True, which_stage=None):
    if which_stage is None:
        which_stage = [1, 2, 3, 4]
    if mage is None:
        return None
    mage, stage = mage.copy(), stage.copy()
    is_stage = np.zeros_like(stage, dtype=bool)
    for s in which_stage:
        is_stage |= (stage == s)
    if np.sum(stage > 0) == 0:
        return None
    onset_idx = np.argwhere(stage > 0)[0][0]
    end_idx   = min(len(stage), onset_idx + 2 * minutes)
    mage[:, ~is_stage] = np.nan
    if mean:
        return np.nanmean(mage[:, onset_idx:end_idx], 1)
    return mage[:, onset_idx:end_idx]

def get_mage_stage(filename, dataset, fold=0):
    filename = filename.split('/')[-1]
    STAGE_PREFIX = f'/data/netmit/wifall/ADetect/data/{dataset}/stage/'
    MAGE_PREFIX  = f'/data/netmit/sleep_lab/filtered/MAGE/{dataset}_new/mage/cv_{fold}/'
    GT_PATH      = f'/data/netmit/sleep_lab/filtered/MAGE/{dataset}/c4_m1_multitaper'
    if not os.path.exists(GT_PATH):
        GT_PATH = GT_PATH.replace(f'/{dataset}/', f'/{dataset}_new/')
    if not os.path.exists(GT_PATH):
        return None, None, None
    try:
        mage_gt = np.load(os.path.join(GT_PATH, filename))['data']
        stage   = np.load(os.path.join(STAGE_PREFIX, filename))
        stage   = stage['data'][::int(30 * stage['fs'])]
        stage   = process_stages(stage)
        mage    = np.load(os.path.join(MAGE_PREFIX, filename))['pred']
    except (FileNotFoundError, KeyError):
        return None, None, None
    if mage.shape[1] < 4 * 60 * 2 or len(stage) < 4 * 60 * 2:
        return None, None, None
    if mage.shape[1] != len(stage):
        if len(stage) < mage.shape[1]:
            return None, None, None
    stage = stage[:mage.shape[1]]
    return mage, mage_gt[:, :mage.shape[1]], stage

def mean_percent_difference(observed, expected):
    b75 = np.nanpercentile(expected, 75, 0)
    b25 = np.nanpercentile(expected, 25, 0)
    b_iqr = 1.5 * (b75 - b25)
    bmin  = b25 - 1.5 * b_iqr
    bmax  = b75 + 1.5 * b_iqr
    a = (observed - bmin) / (bmax - bmin)
    b = (expected - bmin) / (bmax - bmin)
    return (np.nanmean(a, 0) - np.nanmean(b, 0)) / np.nanmean(b, 0) * 100

def bootstrap_percent_difference(observed, expected, n_bootstrap=1000, ci=95):
    n_time = observed.shape[1]
    diffs  = np.zeros((n_bootstrap, n_time))
    for i in tqdm(range(n_bootstrap), desc='Bootstrap 6c'):
        obs = observed[np.random.choice(observed.shape[0], observed.shape[0], replace=True)]
        exp = expected[np.random.choice(expected.shape[0], expected.shape[0], replace=True)]
        diffs[i] = mean_percent_difference(obs, exp)
    lower = np.percentile(diffs, (100 - ci) / 2, axis=0)
    upper = np.percentile(diffs, 100 - (100 - ci) / 2, axis=0)
    return mean_percent_difference(observed, expected), lower, upper

def compute_6c():
    precomputed = os.path.join(PRECOMPUTED_DIR, 'figure_6c_power.csv')
    if os.path.exists(precomputed):
        print('[6c] Loading pre-computed power data...')
        df = pd.read_csv(precomputed)
        return df['pct_diff'].values, df['lower_ci'].values, df['upper_ci'].values

    datasets = ['mros', 'wsc', 'shhs', 'cfs']
    df = pd.read_csv(os.path.join(DATA_DIR, 'master_dataset.csv'))
    df = df[['filename', 'label', 'dataset', 'mit_age', 'mit_gender', 'fold']]
    df = df[df['dataset'].isin(datasets)]
    df = df.groupby('filename').agg('first').reset_index()

    all_controls = df[df['label'] == 0]['filename'].tolist()
    all_antideps = df[df['label'] == 1]['filename'].tolist()

    antidep_pwr, control_pwr = [], []

    for file in tqdm(all_antideps, desc='Antidep MAGE'):
        dataset = get_dataset(file)
        fold = int(df[df['filename'] == file]['fold'].values[0]) if dataset != 'wsc' else random.randint(0, 3)
        mg, _, st = get_mage_stage(file, dataset, fold)
        if mg is None: continue
        pwr = naive_power_post_onset(mg, st)
        if pwr is not None and not np.any(np.isnan(pwr)) and not np.any(np.isinf(pwr)):
            antidep_pwr.append(pwr)

    for file in tqdm(all_controls, desc='Control MAGE'):
        dataset = get_dataset(file)
        fold = int(df[df['filename'] == file]['fold'].values[0]) if dataset != 'wsc' else random.randint(0, 3)
        mg, _, st = get_mage_stage(file, dataset, fold)
        if mg is None: continue
        pwr = naive_power_post_onset(mg, st)
        if pwr is not None and not np.any(np.isnan(pwr)) and not np.any(np.isinf(pwr)):
            control_pwr.append(pwr)

    antidep_pwr = np.stack(antidep_pwr)
    control_pwr = np.stack(control_pwr)
    mean_diff, lower, upper = bootstrap_percent_difference(antidep_pwr, control_pwr)
    return mean_diff, lower, upper


# ── Combined figure ───────────────────────────────────────────────────────────

def generate_figure_6(save=True):
    # ── compute data ──
    print('Computing figure 6a/b...')
    x_tsne, y_lat, y_score, y_label, y_lat_all, y_score_all = compute_6ab(fold=0)

    print('Computing figure 6c...')
    mean_diff, lower, upper = compute_6c()

    # ── layout ──
    fig = plt.figure(figsize=(6.5, 10))
    gs  = gridspec.GridSpec(3, 2, figure=fig, height_ratios=[1, 1, 1],
                            hspace=0.2, wspace=0.3)
    ax_6a  = fig.add_subplot(gs[0, 0])
    ax_6b  = fig.add_subplot(gs[0, 1])
    ax_box = fig.add_subplot(gs[1, :])
    ax_6c  = fig.add_subplot(gs[2, :])

    # ── 6a: t-SNE coloured by model score ──
    tsne_progression_plot(x_tsne, y_score.values, ax=ax_6a, method='points',
                          cmap='magma', s=8, point_frac=1.0, alpha=0.6,
                          overlay_points_frac=0.0, label='Score')
    ax_6a.set_facecolor('dimgray')
    ax_6a.set_title('TSNE Embedding Representation\nColored by Model Score', fontsize=FONT_SIZE)

    # ── 6b: t-SNE coloured by REM latency ──
    tsne_progression_plot(x_tsne, y_lat.values, ax=ax_6b, method='points',
                          cmap='magma', s=8, point_frac=1.0, alpha=0.6,
                          overlay_points_frac=0.0, label='Minutes')
    ax_6b.set_facecolor('dimgray')
    ax_6b.set_title('TSNE Embedding Representation\nColored by REM Latency', fontsize=FONT_SIZE)

    # ── boxplot: REM latency vs model score ──
    plot_binned_boxplots(y_lat_all.values, y_score_all.values, 30, 240, ax=ax_box)
    ax_box.tick_params(axis='both', labelsize=FONT_SIZE)

    # ── 6c: EEG power spectrum ──
    freq = np.linspace(0, 32, 256)
    ax_6c.plot(freq, smooth(mean_diff, 3), ls='dashed', color='black', label='Sleep')
    ax_6c.fill_between(freq, smooth(lower, 3), smooth(upper, 3), alpha=0.1, color='black')
    ax_6c.axhline(y=0, color='gray', ls='dotted', alpha=0.8)
    for f in [1, 4, 8, 12, 16]:
        ax_6c.axvline(f, color='gray', ls='dotted', alpha=0.5)
    ax_6c.set_xlabel('Frequency (Hz)', fontsize=FONT_SIZE)
    ax_6c.set_ylabel('% Difference in Power\n(Antidep − Control)', fontsize=FONT_SIZE)
    ax_6c.set_xlim(0, 32)
    ax_6c.set_xticks(np.arange(0, 33, 4))
    ax_6c.set_ylim(-5, 25)
    ax_6c.tick_params(axis='both', labelsize=FONT_SIZE)

    if save:
        out = os.path.join(OUT_DIR, 'figure_6.pdf')
        plt.savefig(out, dpi=300, bbox_inches='tight')
        print(f'Saved {out}')

    write_source_data_xlsx(x_tsne, y_score.values, y_lat.values, y_label.values,
                           y_lat_all.values, y_score_all.values,
                           mean_diff, lower, upper)
    return fig


# ── XLSX export ───────────────────────────────────────────────────────────────

def write_source_data_xlsx(x_tsne, score, latency_half, label,
                            latency_all, score_all,
                            mean_diff, lower, upper):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    SUB_FILL    = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    SUB_FONT    = Font(bold=True)

    def write_header(ws, col, title, sublabels):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        if len(sublabels) > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + len(sublabels) - 1)
        for j, lbl in enumerate(sublabels):
            c = ws.cell(row=2, column=col + j, value=lbl)
            c.font = SUB_FONT; c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col + j)].width = 16

    wb = Workbook()

    # ── Sheet 1: 6a (tSNE Model Score) ──
    ws1 = wb.active
    ws1.title = '6a (tSNE Score)'
    write_header(ws1, 1, 'tSNE Model Score', ['tSNE Dim1', 'tSNE Dim2', 'Model Score', 'Label'])
    for i in range(len(score)):
        ws1.cell(row=3 + i, column=1, value=round(float(x_tsne[i, 0]), 4))
        ws1.cell(row=3 + i, column=2, value=round(float(x_tsne[i, 1]), 4))
        ws1.cell(row=3 + i, column=3, value=round(float(score[i]), 4))
        ws1.cell(row=3 + i, column=4, value=int(label[i]))

    # ── Sheet 2: 6b (tSNE REM Latency) ──
    ws2 = wb.create_sheet('6b (tSNE REM Latency)')
    write_header(ws2, 1, 'tSNE REM Latency', ['tSNE Dim1', 'tSNE Dim2', 'REM Latency (min)', 'Label'])
    for i in range(len(latency_half)):
        ws2.cell(row=3 + i, column=1, value=round(float(x_tsne[i, 0]), 4))
        ws2.cell(row=3 + i, column=2, value=round(float(x_tsne[i, 1]), 4))
        ws2.cell(row=3 + i, column=3, value=round(float(latency_half[i]), 2))
        ws2.cell(row=3 + i, column=4, value=int(label[i]))

    # ── Sheet 3: 6ab boxplot (raw data) ──
    ws3 = wb.create_sheet('6ab (Boxplot Data)')
    write_header(ws3, 1, 'REM Latency vs Model Score', ['REM Latency (min)', 'Model Score'])
    for i in range(len(latency_all)):
        ws3.cell(row=3 + i, column=1, value=round(float(latency_all[i]), 2))
        ws3.cell(row=3 + i, column=2, value=round(float(score_all[i]), 4))

    # ── Sheet 4: 6c (EEG Power Spectrum) ──
    ws4 = wb.create_sheet('6c (EEG Power Spectrum)')
    write_header(ws4, 1, 'EEG Power % Difference (Antidep - Control)',
                 ['Frequency (Hz)', '% Difference', 'Lower 95% CI', 'Upper 95% CI'])
    freq = np.linspace(0, 32, 256)
    for i in range(256):
        ws4.cell(row=3 + i, column=1, value=round(float(freq[i]), 3))
        ws4.cell(row=3 + i, column=2, value=round(float(mean_diff[i]), 4))
        ws4.cell(row=3 + i, column=3, value=round(float(lower[i]), 4))
        ws4.cell(row=3 + i, column=4, value=round(float(upper[i]), 4))

    out = os.path.join(OUT_DIR, 'figure_6_source_data.xlsx')
    wb.save(out)
    print(f'Saved {out}')


if __name__ == '__main__':
    generate_figure_6(save=True)
