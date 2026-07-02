import os
import numpy as np
import pandas as pd
from sklearn.metrics import roc_auc_score
import seaborn as sns
import matplotlib.pyplot as plt
from tqdm import tqdm

CSV_DIR = 'data/'
INFERENCE_FILE = os.path.join(CSV_DIR, 'inference_v6emb_3920_all.csv')
font_size = 13


def bootstrap_auroc_ci(y_true, y_prob, n_bootstrap=1000, ci=0.95, random_state=42):
    rng = np.random.RandomState(random_state)
    bootstrapped_scores = []

    y_true = np.array(y_true)
    y_prob = np.array(y_prob)

    for _ in tqdm(range(n_bootstrap), desc="Bootstrap sampling"):
        indices = rng.choice(np.arange(len(y_true)), size=len(y_true), replace=True)
        if len(np.unique(y_true[indices])) < 2:
            continue
        score = roc_auc_score(y_true[indices], y_prob[indices])
        bootstrapped_scores.append(score)

    sorted_scores = np.sort(bootstrapped_scores)
    lower_idx = int((1.0 - ci) / 2.0 * len(sorted_scores))
    upper_idx = int((1.0 + ci) / 2.0 * len(sorted_scores))
    lower = sorted_scores[lower_idx]
    upper = sorted_scores[upper_idx]
    mean = np.mean(bootstrapped_scores)

    print(f"AUROC: {mean:.4f} (95% CI: {lower:.4f} - {upper:.4f})")
    return mean, lower, upper


def load_data():
    df = pd.read_csv(INFERENCE_FILE)
    df['filename'] = df['filepath'].apply(lambda x: x.split('/')[-1])
    df = df.groupby('filename').agg(
        lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[0]
    )
    df['pred'] = 1 / (1 + np.exp(-df['pred']))
    df['pid'] = df.apply(
        lambda x: x['pid'][1:] if x['dataset'] in ['shhs', 'mros', 'wsc'] else x['pid'],
        axis=1
    )
    df = df.groupby(['pid', 'label'], as_index=False).agg(
        lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[0]
    )
    return df


def generate_age_gender_figure(save=True):
    df = load_data()

    df['age_bin'] = pd.cut(df['mit_age'], bins=range(0, 100, 10), right=False)

    age_results = {}
    gender_results = {}

    df_age = df[df['mit_age'].notna()].copy()
    for age_bin in df_age['age_bin'].unique():
        df_bin = df_age[df_age['age_bin'] == age_bin]
        if len(df_bin) == 0:
            continue
        auroc, lower, upper = bootstrap_auroc_ci(df_bin['label'].values, df_bin['pred'].values)
        age_results[age_bin] = (auroc, lower, upper)

    df_gender = df[df['mit_gender'].notna()].copy()
    for gender in df_gender['mit_gender'].unique():
        df_bin = df_gender[df_gender['mit_gender'] == gender]
        if len(df_bin) == 0:
            continue
        auroc, lower, upper = bootstrap_auroc_ci(df_bin['label'].values, df_bin['pred'].values)
        gender_results[gender] = (auroc, lower, upper)

    # Print results
    age_bins_sorted = sorted(age_results.keys(), key=lambda x: x.left)
    for age_bin in age_bins_sorted:
        auroc, lower, upper = age_results[age_bin]
        print(f"{int(age_bin.left)}-{int(age_bin.right)}: {auroc:.2f} [{lower:.2f}-{upper:.2f}]")
    for gender in sorted(gender_results.keys()):
        auroc, lower, upper = gender_results[gender]
        label = ['', 'Male', 'Female'][int(gender)]
        print(f"{label}: {auroc:.2f} [{lower:.2f}-{upper:.2f}]")

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

    # Age plot
    age_labels = [f"{int(b.left)}-{int(b.right)}" for b in age_bins_sorted]
    sns.barplot(x=age_labels, y=[age_results[b][0] for b in age_bins_sorted], ax=ax1, palette='Greens')
    for i, age_bin in enumerate(age_bins_sorted):
        auroc, lower, upper = age_results[age_bin]
        ax1.errorbar(i, auroc, yerr=[[auroc - lower], [upper - auroc]], fmt='none', ecolor='black', capsize=5)
        n = len(df_age[df_age['age_bin'] == age_bin])
        ax1.text(i, upper + 0.01, f'N={n}', ha='center', va='bottom', fontsize=font_size - 2)
    ax1.set_ylim(0, 1)
    ax1.set_yticks([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0])
    ax1.set_xlabel('Age Range (years)', fontsize=font_size)
    ax1.set_ylabel('AUROC', fontsize=font_size)
    ax1.set_title('Model Performance by Age Group', fontsize=font_size)
    ax1.tick_params(axis='both', labelsize=font_size)

    # Gender plot
    genders_sorted = sorted(gender_results.keys())
    gender_labels = [['', 'Male', 'Female'][int(g)] for g in genders_sorted]
    sns.barplot(x=gender_labels, y=[gender_results[g][0] for g in genders_sorted], ax=ax2, width=0.3, palette='Greens')
    for i, gender in enumerate(genders_sorted):
        auroc, lower, upper = gender_results[gender]
        ax2.errorbar(i, auroc, yerr=[[auroc - lower], [upper - auroc]], fmt='none', ecolor='black', capsize=5)
        n = len(df_gender[df_gender['mit_gender'] == gender])
        ax2.text(i, upper + 0.01, f'N={n}', ha='center', va='bottom', fontsize=font_size - 2)
    ax2.set_ylim(0, 1)
    ax2.set_yticks([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0])
    ax2.set_xlabel('Sex', fontsize=font_size)
    ax2.set_ylabel('AUROC', fontsize=font_size)
    ax2.set_title('Model Performance by Sex', fontsize=font_size)
    ax2.set_xlim(-0.5, len(genders_sorted) - 0.5)
    ax2.tick_params(axis='both', labelsize=font_size)

    if save:
        plt.tight_layout()
        plt.savefig(os.path.join(os.path.dirname(__file__), 'figure_3d.png'), dpi=300, bbox_inches='tight')

    write_source_data_xlsx(df_age, df_gender, age_results, gender_results, age_bins_sorted, genders_sorted)

    return ax1, ax2


def write_source_data_xlsx(df_age, df_gender, age_results, gender_results, age_bins_sorted, genders_sorted):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUB_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUB_FONT    = Font(bold=True)

    wb = Workbook()

    # ── Sheet 1: Age Analysis ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Age Analysis"

    col_start = 1
    for age_bin in age_bins_sorted:
        label = f"{int(age_bin.left)}-{int(age_bin.right)}"
        ctrl = df_age[(df_age['age_bin'] == age_bin) & (df_age['label'] == 0)]['pred'].round(3).values
        antd = df_age[(df_age['age_bin'] == age_bin) & (df_age['label'] == 1)]['pred'].round(3).values

        if len(ctrl) == 0 and len(antd) == 0:
            continue

        cell = ws1.cell(row=1, column=col_start, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws1.merge_cells(start_row=1, start_column=col_start,
                        end_row=1,   end_column=col_start + 1)

        for j, sublabel in enumerate(["Control", "Antidepressant"]):
            c = ws1.cell(row=2, column=col_start + j, value=sublabel)
            c.font = SUB_FONT
            c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')

        for ri in range(max(len(ctrl), len(antd))):
            for j, arr in enumerate([ctrl, antd]):
                ws1.cell(row=3 + ri, column=col_start + j,
                         value=float(arr[ri]) if ri < len(arr) else None)

        for j in range(2):
            ws1.column_dimensions[get_column_letter(col_start + j)].width = 18
        col_start += 3

    # ── Sheet 2: Gender Analysis ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Gender Analysis")

    col_start = 1
    for gender in genders_sorted:
        label = ['', 'Male', 'Female'][int(gender)]
        ctrl = df_gender[(df_gender['mit_gender'] == gender) & (df_gender['label'] == 0)]['pred'].round(3).values
        antd = df_gender[(df_gender['mit_gender'] == gender) & (df_gender['label'] == 1)]['pred'].round(3).values

        cell = ws2.cell(row=1, column=col_start, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws2.merge_cells(start_row=1, start_column=col_start,
                        end_row=1,   end_column=col_start + 1)

        for j, sublabel in enumerate(["Control", "Antidepressant"]):
            c = ws2.cell(row=2, column=col_start + j, value=sublabel)
            c.font = SUB_FONT
            c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')

        for ri in range(max(len(ctrl), len(antd))):
            for j, arr in enumerate([ctrl, antd]):
                ws2.cell(row=3 + ri, column=col_start + j,
                         value=float(arr[ri]) if ri < len(arr) else None)

        for j in range(2):
            ws2.column_dimensions[get_column_letter(col_start + j)].width = 18
        col_start += 3

    out_path = os.path.join(os.path.dirname(__file__), 'figure_3d_source_data.xlsx')
    wb.save(out_path)
    print(f'Saved {out_path}')


if __name__ == '__main__':
    generate_age_gender_figure(save=True)
