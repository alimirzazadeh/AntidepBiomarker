import pandas as pd
import numpy as np
from ipdb import set_trace as bp
import seaborn as sns
import matplotlib.pyplot as plt
from scipy.stats import ttest_ind as _ttest_ind_base
def ttest_ind(a, b, equal_var=False, **kwargs):  # Welch's by default
    return _ttest_ind_base(a, b, equal_var=equal_var, **kwargs)
import matplotlib.dates as mdates
from figure_5 import apply_smoothing_filter, extract_patient_cohorts, load_and_preprocess_data
FONT_SIZE = 12
BRACKET_COLOR = '#666666'
PLOT_TYPE = 'box'   # 'box' or 'violin'
USE_BEESWARM = False  # overlay beeswarm on box or violin when True

def get_significance_stars(pval):
    if pval < 0.001:
        return '***'
    elif pval < 0.01:
        return '**'
    elif pval < 0.05:
        return '*'
    else:
        return 'ns'


def write_source_data_xlsx(cohorts, cohort_titles, section0, section1_orig, section1_ext, section2, section3):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUB_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUB_FONT    = Font(bold=True)

    wb = Workbook()

    # Panels A-D: patient trajectory subplots (reading order: top-left, top-right, bot-left, bot-right)
    # plot_order=[1,0,3,2] maps cohort indices to axes in the figure
    panel_cohort_order = [1, 3, 0, 2]
    panel_letters = ['A', 'B', 'C', 'D']

    for sheet_idx, (cohort_idx, letter) in enumerate(zip(panel_cohort_order, panel_letters)):
        df_c = cohorts[cohort_idx].copy().sort_values(by='date')
        title = cohort_titles[cohort_idx]
        smoothed = apply_smoothing_filter(df_c['pred'].values)

        ws = wb.active if sheet_idx == 0 else wb.create_sheet()
        ws.title = f'Panel {letter} - {title}'[:31]

        hdr = ws.cell(row=1, column=1, value=f'Panel {letter}: {title}')
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        for j, col_name in enumerate(['Date', 'Model Score (Raw)', 'Model Score (Smoothed)'], 1):
            c = ws.cell(row=2, column=j, value=col_name)
            c.font = SUB_FONT
            c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')

        for ri, (date_val, raw_val, smooth_val) in enumerate(
            zip(df_c['date'].values, df_c['pred'].values, smoothed)
        ):
            ws.cell(row=3 + ri, column=1, value=str(pd.Timestamp(date_val).date()))
            ws.cell(row=3 + ri, column=2, value=round(float(raw_val), 4))
            ws.cell(row=3 + ri, column=3, value=round(float(smooth_val), 4))

        for j in range(1, 4):
            ws.column_dimensions[get_column_letter(j)].width = 24

    # Panel E: Dose Titration scatter (Patient 1007)
    ws_e = wb.create_sheet('Panel E - Dose Titration')
    dose_sections = [
        ('0mg (Control)', section0),
        ('37.5mg', section1_orig),
        ('75mg', section2),
        ('150mg', section3),
    ]
    hdr = ws_e.cell(row=1, column=1,
                    value='Panel E: Venlafaxine Dose Titration (Patient 1007)')
    hdr.font = HEADER_FONT
    hdr.fill = HEADER_FILL
    ws_e.merge_cells(start_row=1, start_column=1, end_row=1,
                     end_column=len(dose_sections) * 3 - 1)

    col_start = 1
    for label, sec in dose_sections:
        sec_sorted = sec.sort_values(by='date')
        c = ws_e.cell(row=2, column=col_start, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
        ws_e.merge_cells(start_row=2, start_column=col_start,
                         end_row=2, end_column=col_start + 1)
        for j, col_name in enumerate(['Date', 'Model Score'], col_start):
            c = ws_e.cell(row=3, column=j, value=col_name)
            c.font = SUB_FONT
            c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')
        for ri, (date_val, pred_val) in enumerate(
            zip(sec_sorted['date'].values, sec_sorted['pred'].values)
        ):
            ws_e.cell(row=4 + ri, column=col_start,
                      value=str(pd.Timestamp(date_val).date()))
            ws_e.cell(row=4 + ri, column=col_start + 1,
                      value=round(float(pred_val), 4))
        for j in range(col_start, col_start + 2):
            ws_e.column_dimensions[get_column_letter(j)].width = 18
        col_start += 3

    # Panels F-H: t-test p-values
    pvalue_panels = [
        ('F', 'Before: ≤37.5mg After: 75mg',   section1_ext, section2),
        ('G', 'Before: 75mg After: 150mg',       section2,     section3),
        ('H', 'Before: ≤37.5mg After: 150mg',   section1_ext, section3),
    ]
    window_sizes_list = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

    for letter, panel_title, before, after in pvalue_panels:
        ws_p = wb.create_sheet(f'Panel {letter} - t-test')
        hdr = ws_p.cell(row=1, column=1, value=f'Panel {letter}: {panel_title}')
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        ws_p.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        for j, col_name in enumerate(['Window Size (Days)', 'T-test p-value'], 1):
            c = ws_p.cell(row=2, column=j, value=col_name)
            c.font = SUB_FONT
            c.fill = SUB_FILL
            c.alignment = Alignment(horizontal='center')

        before_s = before.copy().sort_values(by='date')
        after_s  = after.copy().sort_values(by='date')
        all_sizes  = [1] + window_sizes_list
        all_pvals  = [1.0]
        for ws_size in window_sizes_list:
            pval = ttest_ind(before_s.tail(ws_size)['pred'],
                             after_s.head(ws_size)['pred']).pvalue
            all_pvals.append(float(pval))

        for ri, (ws_size, pval) in enumerate(zip(all_sizes, all_pvals)):
            ws_p.cell(row=3 + ri, column=1, value=ws_size)
            ws_p.cell(row=3 + ri, column=2, value=round(pval, 6))

        for j in range(1, 3):
            ws_p.column_dimensions[get_column_letter(j)].width = 22

    out_path = 'figure_5_pt2_source_data.xlsx'
    wb.save(out_path)
    print(f'Saved {out_path}')


# key = '0.5h_4h_TRIM_right'
# threshold = 34
key = '0h_4h_TRIM_right'
threshold = 40

df = pd.read_csv('../../data/inference_v6emb_3920_all.csv')
df['date'] = pd.to_datetime(df['date'])
# Convert logits to probabilities
df['pred'] = 1 / (1 + np.exp(-df['pred']))

start1 = pd.Timestamp('2020-12-15') + pd.Timedelta(days=0.5)
start2 = pd.Timestamp('2020-12-22') + pd.Timedelta(days=0.5)
start3 = pd.Timestamp('2021-01-28') + pd.Timedelta(days=0.5)
start = start1 - pd.Timedelta(days=23)
end = start3 + pd.Timedelta(days=45)




def create_patient_trajectory_plot(axs, cohorts, titles, save_path, key=None, FILTERED_THRESHOLD=None, disable_title=False):
    """
    Create a 2x2 subplot showing individual patient trajectories.
    
    Args:
        cohorts (list): List of patient DataFrames
        titles (list): List of subplot titles
        save_path (str): Path to save the figure
    """
    # Configuration for subplot arrangement
    FILTERED_THRESHOLD = round(FILTERED_THRESHOLD, 2)
    n_rows, n_cols = 2, 2
    plot_order = [1, 0, 3, 2]  # Custom ordering for presentation
    
    # Create figure and subplots

    
    # Plot each patient cohort
    for plot_idx, cohort_idx in enumerate(plot_order):
        # Get current cohort data
        df_patient = cohorts[cohort_idx].copy()
        title = titles[cohort_idx]
        
        # Sort by date for proper time series visualization
        df_patient = df_patient.sort_values(by='date')
        
        # Calculate subplot position
        row_idx = plot_idx % n_rows
        col_idx = plot_idx // n_rows
        ax = axs[row_idx][col_idx]
        
        # Plot raw data points
        ax.scatter(
            df_patient['date'], 
            df_patient['pred'], 
            color='#7a9973', 
            alpha=0.2,
            label='Raw predictions'
        )
        
        # Apply smoothing and plot trend line
        smoothed_signal = apply_smoothing_filter(df_patient['pred'].values)
        ax.plot(
            df_patient['date'].values, 
            smoothed_signal, 
            alpha=0.6, 
            color='#7a9973',
            linewidth=2,
            label='Smoothed trend'
        )
        ## add a b c d in the top left corner 
        # text(-0.05, 1.05, , fontsize=14, fontweight='bold', transform=ax.transAxes)
        # ax.text(-0.1, 1.1, ['a)', 'c)', 'b)', 'd)'][plot_idx], transform=ax.transAxes, fontsize=14, fontfamily='Calibri', fontweight='bold', 
        #   verticalalignment='top', horizontalalignment='left')
        # Customize subplot
        ax.set_ylim(0, 1)
        ax.set_ylabel('Model Score', fontsize=FONT_SIZE-2)
        ax.set_xlabel('Time (Month)', fontsize=FONT_SIZE-2)
        ax.set_title(title, fontsize=FONT_SIZE, pad=10)
        
        # Format x-axis with monthly ticks
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m'))
        ## set the font size of the x-axis tick labels to FONT_SIZE
        ax.set_xticklabels(ax.get_xticklabels(), fontsize=FONT_SIZE-2)
        ax.set_yticklabels(ax.get_yticklabels(), fontsize=FONT_SIZE-2)
        
        ## write the month number as the x-axis tick labels
        
        # ax.set_xticklabels([f'{int(tick)}' for tick in ax.get_xticks()])
    
    # Add overall figure labels
    # fig.supxlabel('Date (Point=Night, Tick=Month)', y=0.05, fontsize=12)
    # fig.supxlabel('Time (Month)', y=0.05, fontsize=FONT_SIZE)
    # fig.supylabel('Model Score', x=0.09, fontsize=12)
    
    # Adjust layout and save
    # if not disable_title:
    #     fig.suptitle(key + f" FILTERED_THRESHOLD: {FILTERED_THRESHOLD}")
    # plt.tight_layout()
    # plt.subplots_adjust(hspace=0.35, wspace=0.22)
    # save_path = save_path.replace('.png', f'_{key}_{FILTERED_THRESHOLD}.png')
    # plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    # print(f"Figure saved to: {save_path}")

df1007 = df[df['pid'] == '1007'].copy() 
df1007['filename'] = df1007['filename'].apply(lambda x: x.split('/')[-1])
# bad_sig = pd.read_csv('../data/bad_signal_time_1007.csv')
bad_sig = pd.read_csv('../../data/bad_signal_time_4_patients.csv')
print(df1007.shape)
df1007 = pd.merge(df1007, bad_sig, on='filename', how='inner')
print(df1007.shape)


# threshold = 52

# df1007 = df1007[df1007['1h_4h_TRIM_none'] < threshold]

# df1007 = df1007[df1007['0h_4h_TRIM_right'] < threshold]
df1007 = df1007[df1007[key] < threshold]

section0 = df1007[(df1007['pid'] == '1007') & (df1007['date'] > start) &  (df1007['date'] <= start1)]  # Control
section1 = df1007[(df1007['pid'] == '1007') & (df1007['date'] > start1) & (df1007['date'] <= start2)]   # 37.5mg
section2 = df1007[(df1007['pid'] == '1007') & (df1007['date'] > start2) & (df1007['date'] <= start3)]  # 75mg
section3 = df1007[(df1007['pid'] == '1007') & (df1007['date'] > start3) & (df1007['date'] < end)]     # 150mg
section1_orig = section1.copy()  # save before redefinition below
# section0 = section0[section0['0h_4h_TRIM_right'] < threshold]
# section1 = section1[section1['0h_4h_TRIM_right'] < threshold]
# section2 = section2[section2['0h_4h_TRIM_right'] < threshold]
# section3 = section3[section3['0h_4h_TRIM_right'] < threshold]

from matplotlib.gridspec import GridSpec

fig = plt.figure(figsize=(12, 12))
gs = GridSpec(4, 6, figure=fig, height_ratios=[1, 1, 1, 1])


ax0 = fig.add_subplot(gs[0, 0:3])
ax1 = fig.add_subplot(gs[0, 3:6])
ax2 = fig.add_subplot(gs[1, 0:3])
ax3 = fig.add_subplot(gs[1, 3:6])

df_figure1 = load_and_preprocess_data()
cohorts, titles, FILTERED_THRESHOLD = extract_patient_cohorts(df_figure1, key=key, FILTERED_THRESHOLD=threshold)
create_patient_trajectory_plot(axs=[[ax0, ax1], [ax2, ax3]], cohorts=cohorts, titles=titles, save_path='figure_5_pt2.png', key=key, FILTERED_THRESHOLD=FILTERED_THRESHOLD)

ax_top = fig.add_subplot(gs[2, :])
ax_bottom = [fig.add_subplot(gs[3, 2*j:2*j+2]) for j in range(3)]

ax_top.scatter(section0['date'], section0['pred'], color='green', label='0mg')
ax_top.scatter(section1['date'], section1['pred'], color='orange', label='37.5mg')
ax_top.scatter(section2['date'], section2['pred'], color='red', label='75mg')
ax_top.scatter(section3['date'], section3['pred'], color='maroon', label='150mg')
ax_top.axvline(x=start1, color='orange', linestyle='--', linewidth=1)
ax_top.axvline(x=start2, color='red', linestyle='--', linewidth=1)
ax_top.axvline(x=start3, color='maroon', linestyle='--', linewidth=1)

# ax_top.xaxis.set_minor_locator(mdates.WeekdayLocator(byweekday=0))  # every Monday

week_anchor = start
x0, x1 = ax_top.get_xlim()
tick_nums = []
week_labels = []
k = 0
while True:
    t = week_anchor + pd.Timedelta(days=7 * k)
    tn = mdates.date2num(t)
    if tn > x1:
        break
    if tn >= x0:
        tick_nums.append(tn)
        week_labels.append(str(k))
    k += 1
ax_top.set_xticks(tick_nums)
ax_top.set_xticklabels(week_labels)
ax_top.minorticks_off()
ax_top.set_xlabel('Time (Weeks)')
ax_top.set_ylabel('Model Score')
ax_top.set_ylim(0, 1)
ax_top.legend()
ax_top.set_title('Venlafaxine Dose Titration')
## put an X over the dots where bad_signal_time > 50

## setting section 1 to include controls too
section1 = df1007[(df1007['pid'] == '1007') & (df1007['date'] > start) & (df1007['date'] <= start2)]
cohort_titles = titles  # preserve patient titles before overwrite

titles = ['Before: ≤37.5mg After: 75mg', 'Before: 75mg After: 150mg', 'Before: ≤37.5mg After: 150mg']
# dates = [[start2, start2], [start3,start3], [start2, start3]]
for x, (start, end) in enumerate([(section1, section2), (section2, section3), (section1, section3)]):
    j = x
    i = 0
    start = start.copy().sort_values(by='date')
    end = end.copy().sort_values(by='date')
    window_sizes = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    pvalues = []
    for window_size in window_sizes:
        # subset_start = start[start['date'] > (dates[x] - pd.Timedelta(days=window_size))]
        # subset_end = end[end['date'] < (dates[x] + pd.Timedelta(days=window_size))]
        subset_start = start.tail(window_size)
        subset_end = end.head(window_size)
        pvalues.append(ttest_ind(subset_start['pred'], subset_end['pred']).pvalue)
    window_sizes = [1] + window_sizes
    pvalues = [1] + pvalues
    sns.lineplot(x=window_sizes, y=pvalues, ax=ax_bottom[j], errorbar=None)
    ax_bottom[j].set_xlabel('Days since Dose Transition')
    ax_bottom[j].set_ylabel('T-test Before vs. After (p-value)')
    ax_bottom[j].set_ylim(-0.01, 1)
    ax_bottom[j].set_xlim(1, 15)
    ax_bottom[j].set_title(titles[x])
    ax_bottom[j].axhline(y=0.05, color='red', linestyle='--', linewidth=1)
    ax_bottom[j].axhline(y=0.1, color='orange', linestyle='--', linewidth=1)

fig.tight_layout()
plt.savefig(f'figure_5_pt2.pdf', bbox_inches='tight')
write_source_data_xlsx(cohorts, cohort_titles, section0, section1_orig, section1, section2, section3)
# plt.show()
    
# if False: 
#     window_sizes = [3, 5, 7]
#     num_dates = (end - start).days
#     pvalues = []
#     dates = []
#     window_size = window_sizes[1]
#     for days in range(num_dates):
#         current_date = start + pd.Timedelta(days=days)
#         date0 = current_date - 2 * pd.Timedelta(days=window_size)
#         date1 = current_date - 1 * pd.Timedelta(days=window_size)
#         subset0 = df1007[(df1007['date'] > date0) * (df1007['date'] < date1)]
#         subset1 = df1007[(df1007['date'] > date1) * (df1007['date'] < current_date)]
#         pval1 = ttest_ind(subset0['pred'], subset1['pred'], alternative='less').pvalue
#         pval2 = ttest_ind(subset1['pred'], subset0['pred'], alternative='greater').pvalue
#         pvalues.append(min(pval1, pval2))
#         dates.append(current_date)
#     plt.plot(dates, pvalues)
#     plt.xlabel('Date')
#     plt.ylabel('p-value')
#     plt.axvline(x=start1, color='green', linestyle='--', linewidth=1)
#     plt.axvline(x=start2, color='orange', linestyle='--', linewidth=1)
#     plt.axvline(x=start3, color='maroon', linestyle='--', linewidth=1)
#     plt.axhline(y=0.05, color='red', linestyle='--', linewidth=1)
#     plt.axhline(y=0.1, color='red', linestyle='--', linewidth=1)
#     plt.show()

if False: 
    window_sizes = [3, 6, 7]
    num_dates = (end - start).days
    pvalues05 = []
    pvalues10 = []
    dates = []
    window_size = window_sizes[2]
    for days in range(num_dates):
        current_date = start + pd.Timedelta(days=days)
        date0 = current_date - 2 * pd.Timedelta(days=window_size)
        date1 = current_date - 1 * pd.Timedelta(days=window_size)
        subset0 = df1007[(df1007['date'] > date0) * (df1007['date'] < date1)]
        subset1 = df1007[(df1007['date'] > date1) * (df1007['date'] < current_date)]
        pval1 = ttest_ind(subset0['pred'], subset1['pred'], alternative='less').pvalue
        pval2 = ttest_ind(subset1['pred'], subset0['pred'], alternative='greater').pvalue
        
        if pval1 < 0.05:
            pvalues05.append(current_date)
        if pval1 < 0.1:
            pvalues10.append(current_date)
    
    for date in pvalues05:
        ax.scatter(date, 0.05, color='red', marker='x', s=50, lw=2)
    # for date in pvalues10:
    #     ax.scatter(date, 0.1, color='red', marker='x', s=20)
    plt.show()
    
if False:
    def sliding_window_stats(y, N):
        """
        Trailing (causal) window stats:
        each value at time t uses y[t-N+1 : t+1]

        Returns:
            t_vals: time indices (aligned to last day of window)
            means: trailing window means
            stds: trailing window stds
        """
        y = np.asarray(y)
        
        means = []
        stds = []
        
        for t in range(N - 1, len(y)):
            window = y[t - N + 1 : t + 1]
            means.append(window.mean())
            stds.append(window.std())
        
        means = np.array(means)
        stds = np.array(stds)
        
        t_vals = np.arange(N - 1, len(y))  # aligned to last day
        
        return t_vals, means, stds
    def plot_with_error(dates, y, N=5):
        y = np.asarray(y)
        dates = np.array(dates)
        sort_idx = np.argsort(dates)
        dates = dates[sort_idx]
        y = y[sort_idx]
        t = np.arange(len(dates))
        
        centers, means, sigma = sliding_window_stats(y, N)
        
        # plt.figure(figsize=(8,4))
        
        # raw predictions
        # plt.plot(dates, y, 'o-', alpha=0.4, label='Raw predictions')
        
        # smoothed curve
        plt.plot(dates[centers], means, '-', linewidth=2, label=f'Smoothed (N={N})')
        
        # error band (constant width = sigma)
        plt.fill_between(
            dates[centers],
            means - sigma,
            means + sigma,
            alpha=0.2,
            label='±1 std (window variability)'
        )
        
        plt.xlabel("Day")
        plt.ylabel("Prediction score")
        plt.legend()
        plt.tight_layout()
        plt.show()
    segment = (df1007['date'] > start) & (df1007['date'] < end)
    plot_with_error(df1007['date'][segment], df1007['pred'][segment], 7)

if False:
    import numpy as np
    import matplotlib.pyplot as plt
    from sklearn.gaussian_process import GaussianProcessRegressor
    from sklearn.gaussian_process.kernels import RBF, WhiteKernel

    def gp_smooth(dates, pred, length_scale=10.0, noise_level=1.0, confidence=0.95):
        """
        Smooth a 1D prediction series using a Gaussian Process.
    
        Parameters
        ----------
        dates        : array-like of datetime-like or numeric, x-axis values
        pred         : array-like, the raw predictions
        length_scale : float, controls smoothness (higher = smoother)
        noise_level  : float, assumed observation noise variance
        confidence   : float, confidence level for the interval (default 0.95)
    
        Returns
        -------
        x_plot_dates : x values for plotting (same type as dates)
        mu           : posterior mean (smoothed curve)
        lower        : lower confidence bound
        upper        : upper confidence bound
        """
        pred = np.array(pred)
        dates = np.array(dates)
    
        # Convert dates to a float array (nanoseconds → days if datetime, else use as-is)
        if np.issubdtype(dates.dtype, np.datetime64):
            x_num = (dates - dates[0]).astype("timedelta64[s]").astype(float)
        else:
            x_num = dates.astype(float)
    
        X = x_num.reshape(-1, 1)
    
        kernel = RBF(length_scale=length_scale) + WhiteKernel(noise_level=noise_level)
        gp = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=5, normalize_y=True)
        gp.fit(X, pred)
    
        x_plot_num = np.linspace(x_num[0], x_num[-1], 500).reshape(-1, 1)
        mu, sigma = gp.predict(x_plot_num, return_std=True)
    
        from scipy.stats import norm
        z = norm.ppf(1 - (1 - confidence) / 2)
        lower = mu - z * sigma
        upper = mu + z * sigma
    
        # Convert numeric x back to original date type for plotting
        if np.issubdtype(dates.dtype, np.datetime64):
            x_plot_dates = dates[0] + x_plot_num.ravel().astype("timedelta64[s]")
        else:
            x_plot_dates = x_plot_num.ravel()
    
        return x_plot_dates, mu, lower, upper


    def plot_gp_smooth(dates, pred, length_scale=10.0, noise_level=1.0, confidence=0.95):
        pred = np.array(pred)
        dates = np.array(dates)
        x_plot, mu, lower, upper = gp_smooth(dates, pred, length_scale, noise_level, confidence)
    
        fig, ax = plt.subplots(figsize=(12, 5))
    
        # Raw predictions
        ax.scatter(dates, pred, color="steelblue", alpha=0.5,
                s=20, label="Raw predictions", zorder=3)
    
        # Smoothed mean
        ax.plot(x_plot, mu, color="crimson", lw=2, label="GP mean (smoothed)")
    
        # Confidence band
        pct = int(confidence * 100)
        ax.fill_between(x_plot, lower, upper, color="crimson", alpha=0.15,
                        label=f"{pct}% confidence interval")
    
        ax.set_xlabel("Date")
        ax.set_ylabel("Value")
        ax.set_title(f"GP Smoothing  |  length_scale={length_scale}  |  {pct}% CI")
        ax.legend()
        ax.spines[["top", "right"]].set_visible(False)
        fig.autofmt_xdate()
        plt.tight_layout()
        plt.show()
        print("Saved to gp_smooth.png")

    plot_gp_smooth(df1007['date'], df1007['pred'], length_scale=7 * 86_400, noise_level=1.0, confidence=0.95)

if False:
    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    def icc_3_1(x, y):
        """
        ICC(3,1) for two repeated measurements
        x, y: arrays of length N
        """
        x = np.asarray(x)
        y = np.asarray(y)
        
        X = np.vstack([x, y]).T  # shape (N, 2)
        
        n, k = X.shape  # k = 2
        
        mean_per_target = X.mean(axis=1)
        mean_per_rater = X.mean(axis=0)
        grand_mean = X.mean()
        
        # Sum squares
        SS_total = ((X - grand_mean)**2).sum()
        SS_between = k * ((mean_per_target - grand_mean)**2).sum()
        SS_error = SS_total - SS_between
        
        # Mean squares
        MS_between = SS_between / (n - 1)
        MS_error = SS_error / (n * (k - 1))
        
        return (MS_between - MS_error) / (MS_between + (k - 1) * MS_error)
    def sliding_test_retest(dates, preds, N, method="pearson"):
        """
        For each date t with at least 2N observations up to t:
        - recent window  = preds[t-N+1 : t+1]
        - previous window = preds[t-2N+1 : t-N+1]
        Computes correlation between the two adjacent windows.

        Args:
            dates: list-like of dates
            preds: list-like of prediction scores
            N: window size
            method: "pearson" or "spearman"

        Returns:
            DataFrame with columns:
                date
                corr
                prev_mean
                recent_mean
                prev_std
                recent_std
        """
        if N < 2:
            raise ValueError("N must be at least 2 to compute a correlation.")

        df = pd.DataFrame({
            "date": pd.to_datetime(dates),
            "pred": np.asarray(preds, dtype=float)
        }).sort_values("date").reset_index(drop=True)

        corrs = []
        out_dates = []
        prev_means = []
        recent_means = []
        prev_stds = []
        recent_stds = []

        x = df["pred"].to_numpy()
        d = df["date"].to_numpy()

        for t in range(2 * N - 1, len(df)):
            prev_window = x[t - 2 * N + 1 : t - N + 1]
            recent_window = x[t - N + 1 : t + 1]

            # correlation is undefined if either window has zero variance
            if np.std(prev_window) == 0 or np.std(recent_window) == 0:
                corr = np.nan
            else:
                if method == "pearson":
                    corr = np.corrcoef(prev_window, recent_window)[0, 1]
                elif method == "spearman":
                    corr = pd.Series(prev_window).corr(pd.Series(recent_window), method="spearman")
                elif method == "icc":
                    corr = icc_3_1(prev_window, recent_window)
                else:
                    raise ValueError("method must be 'pearson' or 'spearman'")

            out_dates.append(d[t])
            corrs.append(corr)
            prev_means.append(prev_window.mean())
            recent_means.append(recent_window.mean())
            prev_stds.append(prev_window.std())
            recent_stds.append(recent_window.std())

        return pd.DataFrame({
            "date": out_dates,
            "corr": corrs,
            "prev_mean": prev_means,
            "recent_mean": recent_means,
            "prev_std": prev_stds,
            "recent_std": recent_stds,
        })
    segment = (df1007['date'] > start) & (df1007['date'] < end)
    df = sliding_test_retest(df1007['date'][segment], df1007['pred'][segment], 3, method="icc")
    corr = df['corr'].values
    dates = df['date'].values
    sort_idx = np.argsort(dates)
    dates = dates[sort_idx]
    corr = corr[sort_idx]

    plt.plot(dates, corr)
    plt.ylim(-1, 1)
    plt.show()