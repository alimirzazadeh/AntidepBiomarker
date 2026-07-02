import numpy as np 
import pandas as pd 
import os 
from ipdb import set_trace as bp 
import copy 
from tqdm import tqdm 
import seaborn as sns
from scipy.stats import pearsonr
from scipy import stats
import dataframe_image as dfi
import sys 
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
SIMULATED = False
MASTER_DATASET = False

FONT_SIZE=8

if MASTER_DATASET and SIMULATED:
    df = pd.read_csv('../data/SIMULATED_master_dataset.csv')
elif MASTER_DATASET:
    df = pd.read_csv('../data/master_dataset.csv')
elif SIMULATED:
    df = pd.read_csv('SIMULATED_figure_draft_v16_rem_latency.csv')
else:
    df = pd.read_csv('../data/figure_draft_v16_rem_latency.csv')

control_df = df[df['label'] == 0]
antidep_df = df[df['label'] == 1]


def calculate_pval_effect_size(x,y,equal_var=False):
    t,p=stats.ttest_ind(x,y,equal_var=equal_var)
    nx,ny=len(x),len(y)
    md=np.mean(x)-np.mean(y)
    if equal_var:
        sp=np.sqrt(((nx-1)*np.var(x,ddof=1)+(ny-1)*np.var(y,ddof=1))/(nx+ny-2))
    else:
        sp=np.sqrt((np.var(x,ddof=1)+np.var(y,ddof=1))/2)
    return p,md/sp

def compute_boxplot_stats(arr, label):
    """Return a dict of descriptive box-and-whisker statistics for one group."""
    arr = np.array(arr)
    q1  = np.percentile(arr, 25)
    med = np.percentile(arr, 50)
    q3  = np.percentile(arr, 75)
    iqr = q3 - q1
    lower_fence = q1 - 1.5 * iqr
    upper_fence  = q3 + 1.5 * iqr
    lower_whisker = arr[arr >= lower_fence].min()
    upper_whisker = arr[arr <= upper_fence].max()
    return {
        'label':         label,
        'N':             len(arr),
        'min':           arr.min(),
        'max':           arr.max(),
        'lower_whisker': lower_whisker,
        'Q1_25th':       q1,
        'median_50th':   med,
        'Q3_75th':       q3,
        'upper_whisker': upper_whisker,
    }


import matplotlib.pyplot as plt
if True:
    fig, ax = plt.subplots(2, 2, figsize=(6.5, 4))
    figure_3a_stats = []   # accumulate per-subplot stats for the txt file
    figure_3a_raw_data = {}  # raw data points for xlsx export
    for i, name in enumerate(['rem_latency', 'sws_duration', 'rem_duration', 'sleep_efficiency']):
        x1 = control_df[name + '_gt'] / 2
        x2 = control_df[name + '_pred'] / 2
        y1 = antidep_df[name + '_gt'] / 2
        y2 = antidep_df[name + '_pred'] / 2
        if name == 'sleep_efficiency':
            x1 = 2 * x1 
            y1 = 2 * y1 
            x2 = 2 * x2 
            y2 = 2 * y2 
        maskx = ~np.isnan(x1) & ~np.isnan(x2)
        masky = ~np.isnan(y1) & ~np.isnan(y2)
        x1 = x1[maskx]
        x2 = x2[maskx]
        y1 = y1[masky]
        y2 = y2[masky]

        figure_3a_raw_data[name] = {
            'EEG Control':                     np.array(x1),
            'EEG Antidepressant':              np.array(y1),
            'AI (Respiration) Control':        np.array(x2),
            'AI (Respiration) Antidepressant': np.array(y2),
        }

        # Calculate p-values and effect sizes for annotations
        pval_gt, effect_size_gt = calculate_pval_effect_size(x1, y1)
        pval_pred, effect_size_pred = calculate_pval_effect_size(x2, y2)
        print(name, 'GT', pval_gt, effect_size_gt)
        print(name, 'Pred', pval_pred, effect_size_pred)

        # Collect stats for Figure_3a.txt
        figure_3a_stats.append({
            'metric': name,
            'groups': [
                compute_boxplot_stats(x1, 'Expert Annotated (EEG) – Control'),
                compute_boxplot_stats(y1, 'Expert Annotated (EEG) – Antidepressant'),
                compute_boxplot_stats(x2, 'AI Predicted (Respiration) – Control'),
                compute_boxplot_stats(y2, 'AI Predicted (Respiration) – Antidepressant'),
            ],
            'comparisons': [
                {
                    'comparison': 'Expert Annotated (EEG): Control vs Antidepressant',
                    'test':       'Independent samples t-test (Welch\'s; unequal variances assumed)',
                    'sides':      'Two-sided',
                    'adjustment': 'None',
                    'p_value':    pval_gt,
                    'cohens_d':   effect_size_gt,
                },
                {
                    'comparison': 'AI Predicted (Respiration): Control vs Antidepressant',
                    'test':       'Independent samples t-test (Welch\'s; unequal variances assumed)',
                    'sides':      'Two-sided',
                    'adjustment': 'None',
                    'p_value':    pval_pred,
                    'cohens_d':   effect_size_pred,
                },
            ],
        })
        
        # Create data for seaborn boxplot
        data = []
        labels = []
        data.extend(x1)
        labels.extend(['Control\n'] * len(x1))
        data.extend(y1)
        labels.extend(['Antidep\n'] * len(y1))
        data.extend(x2)
        labels.extend(['Control'] * len(x2))
        data.extend(y2)
        labels.extend(['Antidep'] * len(y2))
        
        # Create DataFrame for seaborn
        plot_df = pd.DataFrame({'value': data, 'group': labels})
        
        # Create seaborn boxplot
        sns.boxplot(data=plot_df, x='group', y='value', ax=ax[i // 2, i % 2], showfliers=False, 
                   palette=['royalblue', 'coral', 'royalblue', 'coral'])
        
        legend_patches = [
            mpatches.Patch(color='royalblue', label='Control'),
            mpatches.Patch(color='coral', label='Antidepressant')
        ]
        ## turn off the tick labels 
        ax[i // 2, i % 2].set_xticklabels(['', '', '', ''])
        if i == 3:
            ax[i // 2, i % 2].legend(handles=legend_patches, fontsize=FONT_SIZE, loc='lower right')
            
        # Add text annotations for significance and effect size
        # Function to get significance stars
        def get_significance_stars(pval):
            if pval < .001: #1e-10:
                return '***'
            elif pval < 0.01: #0.001:
                return '**'
            elif pval < 0.05:
                return '*'
            else:
                return 'ns'
        
        # Function to get effect size crosses
        def get_effect_size_crosses(effect_size):
            abs_es = abs(effect_size)
            if abs_es > 0.5:
                return '†††'
            elif abs_es > 0.3:
                return '††'
            elif abs_es > 0.1:
                return '†'
            else:
                return 'ne'
        
        # Get current y limits for positioning
        y_min, y_max = ax[i // 2, i % 2].get_ylim()
        y_pos = y_max - (y_max - y_min) * 0.15  # Position above the boxes
        ax[i//2,i%2].text(0.5,y_min - (y_max - y_min) * 0.25,'Expert Annotated\n(EEG)', ha='center', fontsize=FONT_SIZE)
        ax[i//2,i%2].text(2.5,y_min - (y_max - y_min) * 0.25,'AI Predicted\n(Respiration)', ha='center', fontsize=FONT_SIZE)
        
        # Add annotations for GT comparison (between positions 0 and 1)
        stars_gt = get_significance_stars(pval_gt)
        crosses_gt = get_effect_size_crosses(effect_size_gt)
        annotation_gt = stars_gt + '\n' + crosses_gt if stars_gt or crosses_gt else ''
        
        if annotation_gt:
            # Draw bracket line between GT Control (0) and GT Antidep (1)
            bracket_y = y_pos - (y_max - y_min) * 0.02  # Position bracket slightly below text
            # Offset slightly inward to avoid intersecting with boxplots
            x1_gt = 0.15
            x2_gt = 0.85
            bracket_color = '#666666'  # Gray color matching seaborn boxplot whiskers
            # Draw horizontal line
            ax[i // 2, i % 2].plot([x1_gt, x2_gt], [bracket_y, bracket_y], color=bracket_color, linewidth=1)
            # Draw vertical lines at ends
            ax[i // 2, i % 2].plot([x1_gt, x1_gt], [bracket_y - (y_max - y_min) * 0.01, bracket_y], color=bracket_color, linewidth=1)
            ax[i // 2, i % 2].plot([x2_gt, x2_gt], [bracket_y - (y_max - y_min) * 0.01, bracket_y], color=bracket_color, linewidth=1)
            # Position text above the bracket
            ax[i // 2, i % 2].text(0.5, y_pos, annotation_gt, 
                      ha='center', va='bottom', fontsize=FONT_SIZE, color='black')
        
        # Add annotations for Pred comparison (between positions 2 and 3)
        stars_pred = get_significance_stars(pval_pred)
        crosses_pred = get_effect_size_crosses(effect_size_pred)
        annotation_pred = stars_pred + '\n' + crosses_pred if stars_pred or crosses_pred else ''
        
        if annotation_pred:
            # Draw bracket line between Pred Control (2) and Pred Antidep (3)
            bracket_y = y_pos - (y_max - y_min) * 0.02  # Position bracket slightly below text
            # Offset slightly inward to avoid intersecting with boxplots
            x1_pred = 2.15
            x2_pred = 2.85
            bracket_color = '#666666'  # Gray color matching seaborn boxplot whiskers
            # Draw horizontal line
            ax[i // 2, i % 2].plot([x1_pred, x2_pred], [bracket_y, bracket_y], color=bracket_color, linewidth=1)
            # Draw vertical lines at ends
            ax[i // 2, i % 2].plot([x1_pred, x1_pred], [bracket_y - (y_max - y_min) * 0.01, bracket_y], color=bracket_color, linewidth=1)
            ax[i // 2, i % 2].plot([x2_pred, x2_pred], [bracket_y - (y_max - y_min) * 0.01, bracket_y], color=bracket_color, linewidth=1)
            # Position text above the bracket
            ax[i // 2, i % 2].text(2.5, y_pos, annotation_pred, 
                      ha='center', va='bottom', fontsize=FONT_SIZE, color='black')
        
        # Adjust y limits to accommodate annotations
        ax[i // 2, i % 2].set_ylim(y_min, y_max + (y_max - y_min) * 0.15)
        
        ax[i // 2, i % 2].set_title(name.replace('_', ' ').title().replace('Rem Latency', 'Rapid Eye Momement (REM) Latency').replace('Sws', 'Slow Wave Sleep (SWS)').replace('Rem','REM'), fontsize=FONT_SIZE)
        ax[i // 2, i % 2].set_ylabel('')
        ax[i // 2, i % 2].set_xlabel('')
    ax[0, 0].set_ylabel('Minutes', fontsize=FONT_SIZE)
    ax[1, 0].set_ylabel('Minutes', fontsize=FONT_SIZE)
    ax[0, 1].set_ylabel('Minutes', fontsize=FONT_SIZE)
    ax[1, 1].set_ylabel('Proportion', fontsize=FONT_SIZE)
    ## set xtick and ytick labels to fontsize FONT_SIZE
    ax[0, 0].set_xticklabels(ax[0, 0].get_xticklabels(), fontsize=FONT_SIZE)
    ax[0, 1].set_xticklabels(ax[0, 1].get_xticklabels(), fontsize=FONT_SIZE)
    ax[1, 0].set_xticklabels(ax[1, 0].get_xticklabels(), fontsize=FONT_SIZE)
    ax[1, 1].set_xticklabels(ax[1, 1].get_xticklabels(), fontsize=FONT_SIZE)
    ax[0, 0].set_yticklabels(ax[0, 0].get_yticklabels(), fontsize=FONT_SIZE)
    ax[0, 1].set_yticklabels(ax[0, 1].get_yticklabels(), fontsize=FONT_SIZE)
    ax[1, 0].set_yticklabels(ax[1, 0].get_yticklabels(), fontsize=FONT_SIZE)
    ax[1, 1].set_yticklabels(ax[1, 1].get_yticklabels(), fontsize=FONT_SIZE)
    plt.tight_layout()
    plt.subplots_adjust(hspace=0.36)
    plt.savefig('Figure_3a.pdf')

    # ── Write Figure_3a.txt ──────────────────────────────────────────────────
    with open('Figure_3a.txt', 'w') as f:
        f.write('Figure_3a – Box-plot statistics, statistical tests, and effect sizes\n')
        f.write('=' * 72 + '\n\n')
        f.write('FIGURE DESCRIPTION\n')
        f.write('  2×2 grid of box plots. Each subplot shows four groups:\n')
        f.write('    (1) Expert Annotated (EEG) – Control\n')
        f.write('    (2) Expert Annotated (EEG) – Antidepressant\n')
        f.write('    (3) AI Predicted (Respiration) – Control\n')
        f.write('    (4) AI Predicted (Respiration) – Antidepressant\n')
        f.write('  Outlier fliers are hidden (showfliers=False).\n')
        f.write('  Whiskers extend to the most extreme observed value within\n')
        f.write('  1.5 × IQR of the box edge (Tukey convention).\n\n')

        for entry in figure_3a_stats:
            metric_label = (entry['metric'].replace('_', ' ').title()
                            .replace('Rem Latency', 'REM Latency')
                            .replace('Sws Duration', 'SWS Duration')
                            .replace('Rem Duration', 'REM Duration')
                            .replace('Sleep Efficiency', 'Sleep Efficiency'))
            f.write('─' * 72 + '\n')
            f.write(f'METRIC: {metric_label}  (raw field: {entry["metric"]})\n')
            f.write('─' * 72 + '\n\n')

            f.write('  Box-plot statistics per group\n')
            f.write('  ' + '-' * 68 + '\n')
            header = f'  {"Group":<45} {"N":>5} {"Min":>8} {"Whisker↓":>10} {"Q1(25%)":>9} {"Median":>8} {"Q3(75%)":>9} {"Whisker↑":>10} {"Max":>8}\n'
            f.write(header)
            f.write('  ' + '-' * 68 + '\n')
            for g in entry['groups']:
                f.write(
                    f'  {g["label"]:<45} {g["N"]:>5} '
                    f'{g["min"]:>8.3f} {g["lower_whisker"]:>10.3f} '
                    f'{g["Q1_25th"]:>9.3f} {g["median_50th"]:>8.3f} '
                    f'{g["Q3_75th"]:>9.3f} {g["upper_whisker"]:>10.3f} '
                    f'{g["max"]:>8.3f}\n'
                )
            f.write('\n')

            f.write('  Statistical comparisons\n')
            for cmp in entry['comparisons']:
                f.write('  ' + '-' * 68 + '\n')
                f.write(f'  Comparison  : {cmp["comparison"]}\n')
                f.write(f'  Test        : {cmp["test"]}\n')
                f.write(f'  Sides       : {cmp["sides"]}\n')
                f.write(f'  Adjustment  : {cmp["adjustment"]}\n')
                f.write(f'  p-value     : {cmp["p_value"]:.6e}\n')
                f.write(f'  Effect size : Cohen\'s d = {cmp["cohens_d"]:.4f}  '
                        f'(pooled SD = sqrt(mean of the two sample variances);\n'
                        f'                positive d → first group > second group)\n')
            f.write('\n')

    print('Saved Figure_3a.txt')

## now get the correlation between the ground truth and predicted features 

if True:
    figure_3b_raw_data = {}  # raw data points for xlsx export
    # Create lists to store results
    features = []
    control_correlations = []
    antidep_correlations = []
    control_pvalues = []
    antidep_pvalues = []
    for name in ['rem_latency', 'sws_duration', 'rem_duration', 'sleep_efficiency']:
        x1 = control_df[name + '_gt']
        x2 = control_df[name + '_pred']
        maskx = ~np.isnan(x1) & ~np.isnan(x2) 
        x1 = x1[maskx]
        x2 = x2[maskx]
        y1 = antidep_df[name + '_gt']
        y2 = antidep_df[name + '_pred']
        masky = ~np.isnan(y1) & ~np.isnan(y2)
        y1 = y1[masky]
        y2 = y2[masky]

        figure_3b_raw_data[name] = {
            'ctrl_gt':   np.array(x1),
            'ctrl_pred': np.array(x2),
            'ant_gt':    np.array(y1),
            'ant_pred':  np.array(y2),
        }

        # Calculate correlations
        control_corr = pearsonr(x1, x2)[0]
        antidep_corr = pearsonr(y1, y2)[0]
        control_pvalue = pearsonr(x1, x2)[1]
        antidep_pvalue = pearsonr(y1, y2)[1]
        # Store results
        features.append(name.replace('_', ' ').title().replace('Rem','REM').replace('Sws','SWS'))
        control_correlations.append(control_corr)
        antidep_correlations.append(antidep_corr)
        control_pvalues.append(control_pvalue)
        antidep_pvalues.append(antidep_pvalue)
    # Create DataFrame for nice table display
    
    correlation_df = pd.DataFrame({
        'Feature': features,
        'Pearsons r': [f'{corr:.3f}' for corr in control_correlations],
        'p': [p for p in control_pvalues],
        'Pearsons r ': [f'{corr:.3f}' for corr in antidep_correlations], 
        'p ': [p for p in antidep_pvalues],
    })
    correlation_df.iloc[:,2] = correlation_df.iloc[:,2].apply(lambda x: 'p<1e-10' if x < 1e-10 else f'p={x:.2e}')
    correlation_df.iloc[:,4] = correlation_df.iloc[:,4].apply(lambda x: 'p<1e-10' if x < 1e-10 else f'p={x:.2e}')

    # Display the table
    print("\n" + "="*60)
    print("CORRELATION BETWEEN GROUND TRUTH AND PREDICTED FEATURES")
    print("="*60)
    print("="*15 + "Control" + "="*15 + "Antidep" + "="*15)
    print(correlation_df.to_string(index=False))
    print("="*60)
    
    # Style and export the table as PNG
    styled_df = correlation_df.style.background_gradient(cmap="Blues") \
                         .set_table_styles([
                            {'selector': 'th', 'props': [
                                ('font-size', '12pt'),
                                ('text-align', 'center'),
                                ('min-width', '100px')  # ⬅ widen header cells
                            ]},
                            {'selector': 'td', 'props': [
                                ('font-size', '10pt'),
                                ('text-align', 'center'),
                                ('min-width', '100px')  # ⬅ widen body cells
                            ]},
                            {'selector': 'table', 'props': [
                                ('width', '100%')  # ⬅ make table take full width
                            ]}
                         ]) \
                         .hide(axis="index")  # hide index if not needed

    # Save as PNG
    dfi.export(styled_df, "Figure_3b.png", dpi=600, fontsize=FONT_SIZE) ## increase the resolution

    # ── Write Figure_3b.txt ──────────────────────────────────────────────────
    with open('Figure_3b.txt', 'w') as f:
        f.write('Figure_3b – Pearson correlation statistics (GT vs AI-predicted features)\n')
        f.write('=' * 72 + '\n\n')
        f.write('FIGURE DESCRIPTION\n')
        f.write('  Table of Pearson r between expert-annotated (ground-truth) and\n')
        f.write('  AI-predicted (respiration-derived) sleep features, computed\n')
        f.write('  separately for the Control and Antidepressant groups.\n\n')
        f.write('  Statistical test : Pearson product-moment correlation\n')
        f.write('                     (scipy.stats.pearsonr, two-sided)\n')
        f.write('  Multiple comparisons : No adjustment applied\n\n')

        # Re-derive N values (they are masked above; reuse the per-feature masks)
        f.write('─' * 72 + '\n')
        header = f'  {"Feature":<30} {"Group":<15} {"N":>5} {"Pearson r":>10} {"p-value":>14}\n'
        f.write(header)
        f.write('  ' + '-' * 68 + '\n')

        feature_names_raw = ['rem_latency', 'sws_duration', 'rem_duration', 'sleep_efficiency']
        for feat_raw, feat_label, r_ctrl, p_ctrl, r_ant, p_ant in zip(
            feature_names_raw, features,
            control_correlations, control_pvalues,
            antidep_correlations, antidep_pvalues,
        ):
            # Recompute N using the same masking logic
            xg = control_df[feat_raw + '_gt'];  xp = control_df[feat_raw + '_pred']
            n_ctrl = int((~np.isnan(xg) & ~np.isnan(xp)).sum())
            yg = antidep_df[feat_raw + '_gt']; yp = antidep_df[feat_raw + '_pred']
            n_ant  = int((~np.isnan(yg) & ~np.isnan(yp)).sum())

            f.write(f'  {feat_label:<30} {"Control":<15} {n_ctrl:>5} {r_ctrl:>10.4f} {p_ctrl:>14.6e}\n')
            f.write(f'  {"" :<30} {"Antidepressant":<15} {n_ant:>5} {r_ant:>10.4f} {p_ant:>14.6e}\n')
        f.write('─' * 72 + '\n')

    print('Saved Figure_3b.txt')

# ── Write Figure_3ab_source_data.xlsx ────────────────────────────────────────
_HEADER_FILL    = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT    = Font(bold=True, color="FFFFFF")
_SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SUBHEADER_FONT = Font(bold=True)

wb = Workbook()

# ── Sheet 1: Figure 3a ───────────────────────────────────────────────────────
ws3a = wb.active
ws3a.title = "Figure 3a"

_METRIC_LABELS_3A = {
    'rem_latency':      'Fig 3a (top-left): REM Latency (minutes)',
    'sws_duration':     'Fig 3a (top-right): SWS Duration (minutes)',
    'rem_duration':     'Fig 3a (bottom-left): REM Duration (minutes)',
    'sleep_efficiency': 'Fig 3a (bottom-right): Sleep Efficiency (proportion)',
}
_GROUP_HEADERS_3A = [
    'EEG Control',
    'EEG Antidepressant',
    'AI (Respiration) Control',
    'AI (Respiration) Antidepressant',
]

col_start = 1
for _metric in ['rem_latency', 'sws_duration', 'rem_duration', 'sleep_efficiency']:
    _title = _METRIC_LABELS_3A[_metric]
    _data  = figure_3a_raw_data[_metric]

    # Row 1: title spanning 4 columns
    _cell = ws3a.cell(row=1, column=col_start, value=_title)
    _cell.font = _HEADER_FONT
    _cell.fill = _HEADER_FILL
    _cell.alignment = Alignment(horizontal='center', vertical='center')
    ws3a.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 3)

    # Row 2: group headers
    for _j, _gh in enumerate(_GROUP_HEADERS_3A):
        _c = ws3a.cell(row=2, column=col_start + _j, value=_gh)
        _c.font = _SUBHEADER_FONT
        _c.fill = _SUBHEADER_FILL
        _c.alignment = Alignment(horizontal='center')

    # Rows 3+: individual data points (rounded to 3 decimal places)
    _arrays = [_data[_gh] for _gh in _GROUP_HEADERS_3A]
    _max_n  = max(len(_a) for _a in _arrays)
    for _ri in range(_max_n):
        for _j, _arr in enumerate(_arrays):
            _val = round(float(_arr[_ri]), 3) if _ri < len(_arr) else None
            ws3a.cell(row=3 + _ri, column=col_start + _j, value=_val)

    for _j in range(4):
        ws3a.column_dimensions[get_column_letter(col_start + _j)].width = 30

    col_start += 5  # 4 data cols + 1 blank separator

# ── Sheet 2: Figure 3b ───────────────────────────────────────────────────────
ws3b = wb.create_sheet("Figure 3b")

_METRIC_LABELS_3B = {
    'rem_latency':      'Fig 3b: REM Latency',
    'sws_duration':     'Fig 3b: SWS Duration',
    'rem_duration':     'Fig 3b: REM Duration',
    'sleep_efficiency': 'Fig 3b: Sleep Efficiency',
}
_GROUP_HEADERS_3B = [
    'EEG (GT) – Control',
    'AI Predicted – Control',
    'EEG (GT) – Antidepressant',
    'AI Predicted – Antidepressant',
]

col_start = 1
for _feat in ['rem_latency', 'sws_duration', 'rem_duration', 'sleep_efficiency']:
    _title = _METRIC_LABELS_3B[_feat]
    _fd    = figure_3b_raw_data[_feat]
    _arrays = [_fd['ctrl_gt'], _fd['ctrl_pred'], _fd['ant_gt'], _fd['ant_pred']]

    # Row 1: title spanning 4 columns
    _cell = ws3b.cell(row=1, column=col_start, value=_title)
    _cell.font = _HEADER_FONT
    _cell.fill = _HEADER_FILL
    _cell.alignment = Alignment(horizontal='center', vertical='center')
    ws3b.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 3)

    # Row 2: group headers
    for _j, _gh in enumerate(_GROUP_HEADERS_3B):
        _c = ws3b.cell(row=2, column=col_start + _j, value=_gh)
        _c.font = _SUBHEADER_FONT
        _c.fill = _SUBHEADER_FILL
        _c.alignment = Alignment(horizontal='center')

    # Rows 3+: individual data points (rounded to 3 decimal places)
    _max_n = max(len(_a) for _a in _arrays)
    for _ri in range(_max_n):
        for _j, _arr in enumerate(_arrays):
            _val = round(float(_arr[_ri]), 3) if _ri < len(_arr) else None
            ws3b.cell(row=3 + _ri, column=col_start + _j, value=_val)

    for _j in range(4):
        ws3b.column_dimensions[get_column_letter(col_start + _j)].width = 28

    col_start += 5  # 4 data cols + 1 blank separator

wb.save('Figure_3ab_source_data.xlsx')
print('Saved Figure_3ab_source_data.xlsx')
